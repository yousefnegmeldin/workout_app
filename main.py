from openpyxl import load_workbook
import tkinter as tk
from datetime import date

today = date.today()
d1 = today.strftime("%d/%m/%Y")


upperbody = ['Shoulder Press', 'DB Bench Press', 'Bentover Row', 'Hammer Curl']
lowerbody = ['Squat', 'Leg Press', 'Quad Extension', '', '','']
push1 = ['Shoulder Press', 'Inc DB Bench', 'Chest Fly', 'Tricep Rope', 'Shoulder Raise', 'Skull crusher']
push2 = ['Shoulder Press', 'DB Bench Press', 'Cable Cross', 'Arnold Press', 'Tricep Dips', 'Triecp Rope']
pull1 = ['Bentover Row', 'Pullups', 'Bulkan Row', 'Barbell Shrugs', 'Machine Curl', 'Bar curl']
pull2 = ['Deadlift', 'Pullups', 'Bulkan Row', 'Machine Curl', 'Hammer Curl', 'Face Pulls']
legs1 = ['Stiff Deadlift', 'Leg Press', 'Hamstring Curl', 'Calves', 'Situps' ,'Reverse Crunches']
legs2 = ['Squat', 'Leg Press', 'Quad Extensions', 'Plank', 'Calves', 'Reverse Crunches']

HEIGHT = 600
WIDTH = 700

global index

def after_button_press(day):
    use = ''
    middle_frame.place_forget()
    index = 1
    print(day)
    if day == 'PUSH1':
        use = push1
    elif day == 'PUSH2':
        use = push2
    elif day == 'UPPER':
        use = upperbody
    elif day == 'LOWER':
        use = lowerbody
    elif day == 'PULL1':
        use = pull1
    elif day == 'PULL2':
        use = pull2
    elif day == 'LEGS1':
        use = legs1
    elif day == 'LEGS2':
        use = legs2

    for i in use:
        newtext = 'text' + str(index)
        globals()[newtext] = i
        index = index + 1

    new_frame = tk.Frame(root, bg='#D5D8EC')
    new_frame.place(relwidth=0.34,relheight=0.65,x=25,y=100)

    new_frame2 = tk.Frame(root, bg='#D5D8EC')
    new_frame2.place(relwidth=0.34, relheight=0.65,x=425, y=100)

    toplabel1 = tk.Label(new_frame, text = 'Format ex. 00kg 4x10')
    toplabel1.pack(side='top')

    label1 = tk.Label(new_frame, text = text1)
    label1.place(rely=0.1,relx=0.1)

    label2 = tk.Label(new_frame, text = text2)
    label2.place(rely=0.2, relx=0.1)

    label3 = tk.Label(new_frame, text = text3)
    label3.place(rely=0.3, relx=0.1)

    label4 = tk.Label(new_frame, text = text4)
    label4.place(rely=0.4, relx=0.1)

    label5 = tk.Label(new_frame, text = text5)
    label5.place(rely=0.5, relx=0.1)

    label6 = tk.Label(new_frame, text = text6)
    label6.place(rely=0.6, relx=0.1)

    entry1 = tk.Entry(new_frame)
    entry1.place(rely=0.1, relx=0.5, relwidth=0.4)

    entry2 = tk.Entry(new_frame)
    entry2.place(rely=0.2, relx=0.5, relwidth=0.4)

    entry3 = tk.Entry(new_frame)
    entry3.place(rely=0.3, relx=0.5, relwidth=0.4)

    entry4 = tk.Entry(new_frame)
    entry4.place(rely=0.4, relx=0.5, relwidth=0.4)

    entry5 = tk.Entry(new_frame)
    entry5.place(rely=0.5, relx=0.5, relwidth=0.4)

    entry6 = tk.Entry(new_frame)
    entry6.place(rely=0.6, relx=0.5, relwidth=0.4)

    #newframe2
    toplabel2 = tk.Label(new_frame, text='Next workout')
    toplabel1.pack(side='top')

    label7 = tk.Label(new_frame2, text=text1)
    label7.place(rely=0.1, relx=0.1)

    label8 = tk.Label(new_frame2, text=text2)
    label8.place(rely=0.2, relx=0.1)

    label9 = tk.Label(new_frame2, text=text3)
    label9.place(rely=0.3, relx=0.1)

    label10 = tk.Label(new_frame2, text=text4)
    label10.place(rely=0.4, relx=0.1)

    label11 = tk.Label(new_frame2, text=text5)
    label11.place(rely=0.5, relx=0.1)

    label12 = tk.Label(new_frame2, text=text6)
    label12.place(rely=0.6, relx=0.1)

    label13 = tk.Label(new_frame2,text= 'See Next')
    label13.pack(side='top')


    submit_button = tk.Button(new_frame, text= "Submit", command = lambda: assign_to_excel())
    submit_button.place(relwidth= 0.3, relheight=0.1, rely= 0.75,relx= 0.35)


    def get_new(excelday):
        wb = load_workbook('workout.xlsx')
        ws = wb[excelday]
        maxrow = ws.max_row
        for y in range(2,14):
            newvalue = 'newvalue' + str(y)
            globals()[newvalue] = ws.cell(row=maxrow,column=y).value

        #stopped here
        if newvalue2[0] == 3:
            newvalue2[0] = 4
            print(newvalue2)
        elif newvalue2[0] == 4:
            newvalue3 = newvalue3 + 2
            print(newvalue3)


        label14 = tk.Label(new_frame2, text= newvalue2)
        label14.pack(side='bottom')

    def filter_entry(to_filter):
        charnum = 1
        for char in to_filter:
            charnum = charnum + 1
            if char == 'k':
                charlocation = charnum
                break
        part1 = to_filter[:charlocation-2]
        part2 = to_filter[charlocation+1:]
        print(part1, part2)
        return part1,part2

    def assign_to_excel():
        mylist = []
        mylist.append(d1)
        part2append, part1append = (filter_entry(entry1.get()))
        mylist.append(part1append)
        mylist.append(part2append)
        part4append, part3append = (filter_entry(entry2.get()))
        mylist.append(part3append)
        mylist.append(part4append)
        part6append, part5append = (filter_entry(entry3.get()))
        mylist.append(part5append)
        mylist.append(part6append)
        part8append, part7append = (filter_entry(entry4.get()))
        mylist.append(part7append)
        mylist.append(part8append)
        part10append, part9append = (filter_entry(entry5.get()))
        mylist.append(part9append)
        mylist.append(part10append)
        part12append, part11append = (filter_entry(entry6.get()))
        mylist.append(part11append)
        mylist.append(part12append)
        print(mylist)


        def excel(excelday):
            wb = load_workbook('workout.xlsx')
            ws = wb[excelday]
            maxrow = ws.max_row
            for y in range(1, 14):
                ws.cell(row=maxrow + 1, column=y).value = mylist[y - 1]
            wb.save('workout.xlsx')
        excel(day)
    get_new(day)




#format is 00kg 5x10

root = tk.Tk()
root.title('Workout Planner')

canvas = tk.Canvas(root, height=HEIGHT, width=WIDTH, bg='#F2F3F4')
canvas.pack()

middle_frame = tk.Frame(root,bg='#D5D8DC')
middle_frame.place(relwidth=0.34,relheight=0.65, x=220,y=100)

upper_button = tk.Button(middle_frame,text="UPPER", command=lambda: after_button_press('UPPER'))
upper_button.place(relwidth =0.4,relheight=0.15, relx= 0.3, rely= 0.01, anchor = 'nw')

lower_button = tk.Button(middle_frame, text='LOWER', command=lambda: after_button_press('LOWER') )
lower_button.place(relwidth =0.4,relheight=0.15, relx= 0.3,rely=0.2, anchor = 'nw')

push1_button = tk.Button(middle_frame, text='PUSH1', command=lambda: after_button_press('PUSH1'))
push1_button.place(relwidth =0.4,relheight=0.15, relx= 0.075,rely=0.4, anchor = 'nw')

push2_button = tk.Button(middle_frame, text='PUSH2', command=lambda: after_button_press('PUSH2'))
push2_button.place(relwidth =0.4,relheight=0.15, relx= 0.5,rely=0.4, anchor = 'nw')

pull_button = tk.Button(middle_frame, text='PULL1', command=lambda: after_button_press('PULL1'))
pull_button.place(relwidth =0.4,relheight=0.15, relx= 0.075,rely=0.6, anchor = 'nw')

pull2_button = tk.Button(middle_frame, text='PULL2', command=lambda: after_button_press('PULL2'))
pull2_button.place(relwidth =0.4,relheight=0.15, relx= 0.5,rely=0.6, anchor = 'nw')

legs1_button = tk.Button(middle_frame, text='LEGS1', command=lambda: after_button_press('LEGS1'))
legs1_button.place(relwidth =0.4,relheight=0.15, relx= 0.075,rely=0.8, anchor = 'nw')

legs2_button = tk.Button(middle_frame, text='LEGS2', command=lambda: after_button_press('LEGS2'))
legs2_button.place(relwidth =0.4,relheight=0.15, relx= 0.5,rely=0.8, anchor = 'nw')

root.mainloop()
